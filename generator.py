import io
import random
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, colors
)
from openpyxl.utils import get_column_letter

# ── colours ────────────────────────────────────────────────────────────────
NAVY    = "1F3864"
GOLD    = "C9A94B"
CREAM   = "FFF8E7"
LIGHT   = "EEF2F7"
WHITE   = "FFFFFF"
MID     = "D9DFE8"

def _fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)

def _font(bold=False, size=11, colour="000000", italic=False):
    return Font(bold=bold, size=size, color=colour, italic=italic,
                name="Calibri")

def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _bottom_border():
    s = Side(style="thin", color=GOLD)
    return Border(bottom=s)


# ── fair rotation (no consecutive repeats) ────────────────────────────────
def _build_rotation(people: list[str], num_slots: int) -> list[str]:
    if not people:
        return ["—"] * num_slots
    counts: dict[str, int] = {p: 0 for p in people}
    result = []
    last = None
    for _ in range(num_slots):
        min_count = min(counts.values())
        # prefer candidates who weren't assigned last week
        candidates = [p for p, c in counts.items() if c == min_count and p != last]
        if not candidates:
            # only one person in pool — no choice but to repeat
            candidates = [p for p, c in counts.items() if c == min_count]
        random.shuffle(candidates)
        chosen = candidates[0]
        counts[chosen] += 1
        result.append(chosen)
        last = chosen
    return result


# ── main entry point ───────────────────────────────────────────────────────
def generate_excel(
    church_name: str,
    service_dates: list[date],
    duties: list[dict],          # [{"name": str, "people": [str]}]
    hymns: list[str],
    readings: list[dict],        # [{"ref": str, "text": str}]
    hymns_per_service: int = 3,
    readings_per_service: int = 2,
    prayers: list[dict] = None,  # [{"name": str, "text": str}]
) -> bytes:
    service_dates = sorted(service_dates)
    n = len(service_dates)

    # assign people to duties
    duty_assignments: list[list[str]] = []
    for duty in duties:
        duty_assignments.append(_build_rotation(duty["people"], n))

    if prayers is None:
        prayers = []

    # shuffle & assign hymns / readings (cycle if list is shorter than needed)
    hymn_pool = hymns.copy()
    random.shuffle(hymn_pool)
    reading_pool = readings.copy()
    random.shuffle(reading_pool)

    service_hymns: list[list[str]] = []
    service_readings: list[list[dict]] = []
    hymn_idx = 0
    reading_idx = 0
    for _ in range(n):
        sh = []
        for _ in range(hymns_per_service):
            sh.append(hymn_pool[hymn_idx % len(hymn_pool)])
            hymn_idx += 1
        service_hymns.append(sh)

        sr = []
        for _ in range(readings_per_service):
            sr.append(reading_pool[reading_idx % len(reading_pool)])
            reading_idx += 1
        service_readings.append(sr)

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    _build_home_sheet(wb, church_name, service_dates, duties,
                      duty_assignments, service_hymns, service_readings,
                      hymns_per_service, readings_per_service)

    for i, svc_date in enumerate(service_dates):
        _build_service_sheet(
            wb, church_name, svc_date, duties,
            [duty_assignments[d][i] for d in range(len(duties))],
            service_hymns[i], service_readings[i], prayers,
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── HOME sheet ─────────────────────────────────────────────────────────────
def _build_home_sheet(wb, church_name, service_dates, duties,
                      duty_assignments, service_hymns, service_readings,
                      hymns_per_service, readings_per_service):
    ws = wb.create_sheet("Home", 0)
    ws.sheet_view.showGridLines = False

    # title row
    ws.row_dimensions[1].height = 38
    ws.merge_cells("A1:Z1")
    c = ws["A1"]
    c.value = f"{church_name} — Service Roster"
    c.font = _font(bold=True, size=18, colour=WHITE)
    c.fill = _fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # build header row
    headers = ["Date", "Day"]
    for d in duties:
        if d["name"].strip():
            headers.append(d["name"])
    for i in range(hymns_per_service):
        headers.append(f"Hymn {i+1}")
    for i in range(readings_per_service):
        headers.append(f"Reading {i+1}")

    header_row = 3
    ws.row_dimensions[header_row].height = 22
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = _font(bold=True, size=10, colour=WHITE)
        c.fill = _fill(GOLD)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = _border()

    # data rows
    active_duty_indices = [i for i, d in enumerate(duties) if d["name"].strip()]
    for row_offset, svc_date in enumerate(service_dates):
        row = header_row + 1 + row_offset
        ws.row_dimensions[row].height = 30
        fill = _fill(CREAM) if row_offset % 2 == 0 else _fill(WHITE)

        col = 1
        for val in [svc_date.strftime("%d %b %Y"), svc_date.strftime("%A")]:
            c = ws.cell(row=row, column=col, value=val)
            c.font = _font(bold=(col == 1), size=10)
            c.fill = fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _border()
            col += 1

        for di in active_duty_indices:
            assigned = duty_assignments[di][row_offset]
            c = ws.cell(row=row, column=col, value=assigned)
            c.font = _font(size=10)
            c.fill = fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _border()
            col += 1

        for hymn in service_hymns[row_offset]:
            c = ws.cell(row=row, column=col, value=hymn)
            c.font = _font(size=10, italic=True)
            c.fill = fill
            c.alignment = Alignment(wrap_text=True, vertical="center")
            c.border = _border()
            col += 1

        for reading in service_readings[row_offset]:
            ref = reading["ref"] if isinstance(reading, dict) else reading
            c = ws.cell(row=row, column=col, value=ref)
            c.font = _font(size=10)
            c.fill = fill
            c.alignment = Alignment(wrap_text=True, vertical="center")
            c.border = _border()
            col += 1

    # column widths
    col_widths = [14, 12] + [22] * len(active_duty_indices) \
                 + [32] * hymns_per_service + [45] * readings_per_service
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # freeze panes below header
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


# ── helpers ────────────────────────────────────────────────────────────────
def _section_header(ws, row, title):
    ws.row_dimensions[row].height = 20
    ws.merge_cells(f"A{row}:B{row}")
    c = ws[f"A{row}"]
    c.value = title
    c.font = _font(bold=True, size=12, colour=WHITE)
    c.fill = _fill(NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    return row


def _text_block(ws, row, text, fill_colour=LIGHT, size=11):
    """Write a multi-line text block, one row per line, merged A:B."""
    lines = text.splitlines() if text else [""]
    for line in lines:
        ws.row_dimensions[row].height = 16
        ws.merge_cells(f"A{row}:B{row}")
        c = ws[f"A{row}"]
        c.value = f"  {line}" if line else ""
        c.font = _font(size=size)
        c.fill = _fill(fill_colour)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        row += 1
    return row


# ── individual SERVICE sheet ───────────────────────────────────────────────
def _build_service_sheet(wb, church_name, svc_date, duties,
                         assignments, hymns, readings, prayers):
    sheet_name = svc_date.strftime("%d %b %Y")
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 62

    row = 1

    # church name header
    ws.row_dimensions[row].height = 36
    ws.merge_cells(f"A{row}:B{row}")
    c = ws[f"A{row}"]
    c.value = church_name
    c.font = _font(bold=True, size=16, colour=WHITE)
    c.fill = _fill(NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # date sub-header
    row += 1
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f"A{row}:B{row}")
    c = ws[f"A{row}"]
    c.value = svc_date.strftime("%A, %d %B %Y")
    c.font = _font(bold=True, size=13, colour=WHITE)
    c.fill = _fill(GOLD)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    row += 1  # spacer

    # ── HYMNS ──
    row += 1
    _section_header(ws, row, "HYMNS")
    for i, hymn in enumerate(hymns, 1):
        row += 1
        ws.row_dimensions[row].height = 18
        c1 = ws.cell(row=row, column=1, value=f"  {i}.")
        c1.font = _font(bold=True, size=11)
        c1.fill = _fill(CREAM)
        c1.alignment = Alignment(vertical="center")
        c2 = ws.cell(row=row, column=2, value=hymn)
        c2.font = _font(size=11, italic=True)
        c2.fill = _fill(CREAM)
        c2.alignment = Alignment(vertical="center")

    row += 1  # spacer

    # ── SCRIPTURE READINGS ──
    row += 1
    _section_header(ws, row, "SCRIPTURE READINGS")
    ordinals = ["1st Reading", "2nd Reading", "3rd Reading", "4th Reading"]
    for i, reading in enumerate(readings):
        ref  = reading["ref"]  if isinstance(reading, dict) else reading
        text = reading.get("text", "").strip() if isinstance(reading, dict) else ""
        label = ordinals[i] if i < len(ordinals) else f"Reading {i+1}"
        alt_fill = CREAM if i % 2 == 0 else LIGHT

        # reference row
        row += 1
        ws.row_dimensions[row].height = 20
        c1 = ws.cell(row=row, column=1, value=f"  {label}:")
        c1.font = _font(bold=True, size=11)
        c1.fill = _fill(alt_fill)
        c1.alignment = Alignment(vertical="center")
        c2 = ws.cell(row=row, column=2, value=ref)
        c2.font = _font(bold=True, size=11)
        c2.fill = _fill(alt_fill)
        c2.alignment = Alignment(vertical="center")

        # full scripture text (if provided)
        if text:
            row += 1
            row = _text_block(ws, row, text, fill_colour=alt_fill, size=10)
            row -= 1  # _text_block already advanced past last line

        row += 1  # gap between readings

    row += 1  # spacer before duties

    # ── DUTIES ──
    row += 1
    _section_header(ws, row, "DUTIES")
    for i, duty in enumerate(duties):
        if not duty["name"].strip():
            continue
        assigned = assignments[i]
        row += 1
        ws.row_dimensions[row].height = 18
        fill = _fill(CREAM) if i % 2 == 0 else _fill(WHITE)
        c1 = ws.cell(row=row, column=1, value=f"  {duty['name']}:")
        c1.font = _font(bold=True, size=11)
        c1.fill = fill
        c1.alignment = Alignment(vertical="center")
        c1.border = _bottom_border()
        c2 = ws.cell(row=row, column=2, value=assigned)
        c2.font = _font(size=11)
        c2.fill = fill
        c2.alignment = Alignment(vertical="center")
        c2.border = _bottom_border()

    # ── PRAYERS ──
    if prayers:
        row += 2
        _section_header(ws, row, "PRAYERS")
        for pi, prayer in enumerate(prayers):
            name = prayer.get("name", f"Prayer {pi+1}").strip()
            text = prayer.get("text", "").strip()
            if not name and not text:
                continue
            alt_fill = CREAM if pi % 2 == 0 else LIGHT

            # prayer name row
            row += 1
            ws.row_dimensions[row].height = 20
            ws.merge_cells(f"A{row}:B{row}")
            c = ws[f"A{row}"]
            c.value = f"  {name}"
            c.font = _font(bold=True, size=11, colour=NAVY)
            c.fill = _fill(alt_fill)
            c.alignment = Alignment(vertical="center")

            # prayer text
            if text:
                row += 1
                row = _text_block(ws, row, text, fill_colour=alt_fill, size=10)
                row -= 1

            row += 1  # gap between prayers

    row += 2
    ws.row_dimensions[row].height = 14
    ws.merge_cells(f"A{row}:B{row}")
    c = ws[f"A{row}"]
    c.value = "— prepared by Church Roster Generator —"
    c.font = _font(size=9, colour="999999", italic=True)
    c.alignment = Alignment(horizontal="center")
